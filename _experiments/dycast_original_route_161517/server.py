"""弹幕批量托管后端。

职责：
  1. WS 中继 (ws://localhost:8775)：接收前端 BatchView 每个房间转发的
     "房间信息(JSON对象)" + "消息数组(JSON数组)"，按房间/分类去重写入 SQLite。
  2. HTTP 导出 (http://localhost:8776)：提供分类、干净、可按房间过滤的导出。

设计要点：
  - 每个房间一条 WS 连接；连接第一帧是房间信息对象，后续帧是消息数组。
  - 去重：dedup_key = room_num|method|msg_id（无 msg_id 时为 NULL，互不相等）。
  - 分类：把抖音 method 归类为 chat/gift/member/follow/like/rank/other，便于评估。
  - 导出默认聊天弹幕，列精简：房间号、主播、时间、用户、内容。
"""

from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
import sqlite3
import threading
import time
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Final
from urllib.parse import parse_qs, urlparse

import websockets
from websockets.asyncio.server import ServerConnection

from audio_manager import AudioManager

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger: Final = logging.getLogger("danmu-server")

WS_HOST: Final = "localhost"
WS_PORT: Final = 8775
HTTP_HOST: Final = "localhost"
HTTP_PORT: Final = 8776

DATA_DIR: Final = Path(__file__).resolve().parent / "data"
DB_PATH: Final = DATA_DIR / "danmu.db"

# method -> (分类, 中文名)
_METHOD_CATEGORY: Final[dict[str, tuple[str, str]]] = {
    "WebcastChatMessage": ("chat", "聊天"),
    "WebcastEmojiChatMessage": ("chat", "表情聊天"),
    "WebcastGiftMessage": ("gift", "礼物"),
    "WebcastMemberMessage": ("member", "进入"),
    "WebcastSocialMessage": ("follow", "关注"),
    "WebcastLikeMessage": ("like", "点赞"),
    "WebcastRoomUserSeqMessage": ("rank", "在线榜"),
    "WebcastRoomRankMessage": ("rank", "排行榜"),
    "WebcastRoomStatsMessage": ("stat", "房间统计"),
    "WebcastControlMessage": ("control", "房间控制"),
}

# 分类别名 -> method 列表（导出过滤用）
_CATEGORY_METHODS: Final[dict[str, list[str]]] = {}
for _m, (_cat, _zh) in _METHOD_CATEGORY.items():
    _CATEGORY_METHODS.setdefault(_cat, []).append(_m)


# --------------------------------------------------------------------------- #
# 数据库
# --------------------------------------------------------------------------- #
_write_lock: Final = threading.Lock()
_write_conn: sqlite3.Connection | None = None

# 话术采集管理器(在 main() 中初始化)
_audio: AudioManager | None = None


def _connect(readonly: bool = False) -> sqlite3.Connection:
    if readonly:
        conn = sqlite3.connect(
            f"file:{DB_PATH}?mode=ro", uri=True, check_same_thread=False, timeout=5
        )
    else:
        conn = sqlite3.connect(DB_PATH, check_same_thread=False, timeout=5)
    conn.row_factory = sqlite3.Row
    return conn


def init_db() -> None:
    global _write_conn
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    _write_conn = _connect()
    _write_conn.execute("PRAGMA journal_mode=WAL")
    _write_conn.execute("PRAGMA synchronous=NORMAL")
    _write_conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS rooms (
            room_num   TEXT PRIMARY KEY,
            room_id    TEXT,
            nickname   TEXT,
            title      TEXT,
            first_seen INTEGER,
            last_seen  INTEGER,
            msg_count  INTEGER DEFAULT 0
        );

        CREATE TABLE IF NOT EXISTS messages (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            room_num   TEXT NOT NULL,
            room_id    TEXT,
            msg_id     TEXT,
            method     TEXT,
            category   TEXT,
            user_id    TEXT,
            user_name  TEXT,
            content    TEXT,
            gift_name  TEXT,
            gift_count TEXT,
            ts         INTEGER,
            created_at TEXT,
            dedup_key  TEXT UNIQUE
        );

        CREATE INDEX IF NOT EXISTS idx_msg_room    ON messages(room_num);
        CREATE INDEX IF NOT EXISTS idx_msg_method  ON messages(method);
        CREATE INDEX IF NOT EXISTS idx_msg_cat     ON messages(category);

        CREATE TABLE IF NOT EXISTS transcripts (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            room_num   TEXT NOT NULL,
            text       TEXT,
            start_sec  INTEGER,   -- 段起始秒(相对本轮录制)
            ts         INTEGER,   -- 入库 unix 秒
            created_at TEXT,
            dedup_key  TEXT UNIQUE
        );

        CREATE INDEX IF NOT EXISTS idx_tr_room ON transcripts(room_num);
        """
    )
    _write_conn.commit()
    logger.info("数据库就绪 => %s", DB_PATH)


def upsert_room(info: dict[str, Any]) -> str:
    room_num = str(info.get("roomNum") or "").strip()
    if not room_num:
        return ""
    now = int(time.time())
    assert _write_conn is not None
    with _write_lock:
        _write_conn.execute(
            """
            INSERT INTO rooms (room_num, room_id, nickname, title, first_seen, last_seen)
            VALUES (?, ?, ?, ?, ?, ?)
            ON CONFLICT(room_num) DO UPDATE SET
                room_id   = COALESCE(NULLIF(excluded.room_id, ''), rooms.room_id),
                nickname  = COALESCE(NULLIF(excluded.nickname, ''), rooms.nickname),
                title     = COALESCE(NULLIF(excluded.title, ''), rooms.title),
                last_seen = excluded.last_seen
            """,
            (
                room_num,
                str(info.get("roomId") or ""),
                str(info.get("nickname") or ""),
                str(info.get("title") or ""),
                now,
                now,
            ),
        )
        _write_conn.commit()
    return room_num


def _flatten(msg: dict[str, Any], room_num: str) -> tuple:
    method = str(msg.get("method") or "")
    category, _zh = _METHOD_CATEGORY.get(method, ("other", "其他"))
    user = msg.get("user") or {}
    gift = msg.get("gift") or {}
    msg_id = str(msg.get("id") or "")
    dedup_key = f"{room_num}|{method}|{msg_id}" if msg_id else None
    now = time.time()
    return (
        room_num,
        str(msg.get("roomId") or ""),
        msg_id,
        method,
        category,
        str(user.get("id") or ""),
        str(user.get("name") or ""),
        str(msg.get("content") or ""),
        str(gift.get("name") or ""),
        str(gift.get("count") or ""),
        int(now * 1000),
        datetime.fromtimestamp(now).strftime("%Y-%m-%d %H:%M:%S"),
        dedup_key,
    )


def save_messages(room_num: str, msgs: list[dict[str, Any]]) -> int:
    if not msgs:
        return 0
    rows = [_flatten(m, room_num) for m in msgs]
    assert _write_conn is not None
    with _write_lock:
        before = _write_conn.total_changes
        _write_conn.executemany(
            """
            INSERT OR IGNORE INTO messages
                (room_num, room_id, msg_id, method, category, user_id, user_name,
                 content, gift_name, gift_count, ts, created_at, dedup_key)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            rows,
        )
        # INSERT OR IGNORE 的 executemany rowcount 不可靠，用 total_changes 差值取真实插入数
        inserted = _write_conn.total_changes - before
        _write_conn.execute(
            "UPDATE rooms SET msg_count = msg_count + ?, last_seen = ? WHERE room_num = ?",
            (inserted, int(time.time()), room_num),
        )
        _write_conn.commit()
    return inserted


def save_transcript(room_num: str, text: str, start_sec: int, ts: int) -> bool:
    """写入一段话术转写。线程安全(供转写 executor 线程调用)。返回是否新插入。"""
    text = (text or "").strip()
    if not room_num or not text:
        return False
    created_at = datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")
    # 按 房间|入库时刻|文本 去重，避免重连/重复处理产生重复段
    dedup_key = f"{room_num}|{ts}|{hash(text) & 0xFFFFFFFF:x}"
    assert _write_conn is not None
    with _write_lock:
        before = _write_conn.total_changes
        _write_conn.execute(
            "INSERT OR IGNORE INTO transcripts "
            "(room_num, text, start_sec, ts, created_at, dedup_key) VALUES (?,?,?,?,?,?)",
            (room_num, text, int(start_sec), int(ts), created_at, dedup_key),
        )
        inserted = _write_conn.total_changes - before
        _write_conn.commit()
    return inserted > 0


# --------------------------------------------------------------------------- #
# WS 中继
# --------------------------------------------------------------------------- #
async def relay_handler(ws: ServerConnection) -> None:
    room_num = ""
    saved = 0
    try:
        async for raw in ws:
            try:
                data = json.loads(raw)
            except (json.JSONDecodeError, TypeError):
                continue
            if isinstance(data, dict):
                # 房间信息帧
                room_num = upsert_room(data) or room_num
                logger.info(
                    "房间接入 => %s %s", room_num, data.get("nickname") or data.get("title") or ""
                )
                # 同一开播信号同时拉起话术采集
                if _audio is not None and room_num:
                    live_url = data.get("liveUrl") or data.get("live_url") or None
                    _audio.start_room(room_num, live_url)
            elif isinstance(data, list) and room_num:
                # 消息数组帧
                msgs = [m for m in data if isinstance(m, dict)]
                n = save_messages(room_num, msgs)
                saved += n
                if n:
                    last = next(
                        (m.get("content") for m in reversed(msgs) if m.get("content")), ""
                    )
                    logger.info("房间 %s 入库 %d 条 (累计 %d) 最新: %s", room_num, n, saved, last)
    except websockets.ConnectionClosed:
        pass
    except Exception:  # noqa: BLE001 - 单房间出错不应拖垮服务
        logger.exception("中继处理异常 room=%s", room_num)
    finally:
        if _audio is not None and room_num:
            _audio.stop_room(room_num)
        logger.info("房间断开 => %s 本次累计 %d 条", room_num or "?", saved)


# --------------------------------------------------------------------------- #
# HTTP 导出
# --------------------------------------------------------------------------- #
def _resolve_methods(qs: dict[str, list[str]]) -> list[str] | None:
    """根据 query 返回要过滤的 method 列表；None 表示不过滤。"""
    if "method" in qs and qs["method"][0]:
        return [m for m in qs["method"][0].split(",") if m]
    if "category" in qs and qs["category"][0]:
        methods: list[str] = []
        for cat in qs["category"][0].split(","):
            methods.extend(_CATEGORY_METHODS.get(cat, []))
        return methods or ["__none__"]
    return None


def _query_messages(qs: dict[str, list[str]]) -> list[sqlite3.Row]:
    where: list[str] = []
    params: list[Any] = []
    methods = _resolve_methods(qs)
    if methods is not None:
        where.append(f"m.method IN ({','.join('?' * len(methods))})")
        params.extend(methods)
    if "room" in qs and qs["room"][0]:
        rooms = [r for r in qs["room"][0].split(",") if r]
        where.append(f"m.room_num IN ({','.join('?' * len(rooms))})")
        params.extend(rooms)
    clause = f"WHERE {' AND '.join(where)}" if where else ""
    sql = (
        "SELECT m.room_num, COALESCE(r.nickname,'') AS nickname, m.created_at, "
        "m.method, m.category, m.user_name, m.content, m.gift_name, m.gift_count "
        "FROM messages m LEFT JOIN rooms r ON r.room_num = m.room_num "
        f"{clause} ORDER BY m.room_num, m.id"
    )
    conn = _connect(readonly=True)
    try:
        return conn.execute(sql, params).fetchall()
    finally:
        conn.close()


def _rooms_summary() -> list[sqlite3.Row]:
    conn = _connect(readonly=True)
    try:
        return conn.execute(
            """
            SELECT r.room_num, r.nickname, r.title, r.last_seen,
                   COUNT(m.id)                                            AS total,
                   SUM(CASE WHEN m.category='chat' THEN 1 ELSE 0 END)     AS chat,
                   SUM(CASE WHEN m.category='gift' THEN 1 ELSE 0 END)     AS gift,
                   SUM(CASE WHEN m.category='member' THEN 1 ELSE 0 END)   AS member
            FROM rooms r LEFT JOIN messages m ON m.room_num = r.room_num
            GROUP BY r.room_num ORDER BY total DESC
            """
        ).fetchall()
    finally:
        conn.close()


def _transcript_counts() -> dict[str, tuple[int, int]]:
    """room_num -> (话术段数, 话术总字数)。"""
    conn = _connect(readonly=True)
    try:
        rows = conn.execute(
            "SELECT room_num, COUNT(*) AS segs, "
            "COALESCE(SUM(LENGTH(text)),0) AS chars "
            "FROM transcripts GROUP BY room_num"
        ).fetchall()
    finally:
        conn.close()
    return {r["room_num"]: (r["segs"], r["chars"]) for r in rows}


def _transcripts_for(room_num: str) -> list[sqlite3.Row]:
    conn = _connect(readonly=True)
    try:
        return conn.execute(
            "SELECT created_at, start_sec, text FROM transcripts "
            "WHERE room_num=? ORDER BY id",
            (room_num,),
        ).fetchall()
    finally:
        conn.close()


def _room_nick_map() -> dict[str, str]:
    conn = _connect(readonly=True)
    try:
        rows = conn.execute(
            "SELECT room_num, COALESCE(nickname,'') AS nk FROM rooms"
        ).fetchall()
    finally:
        conn.close()
    return {r["room_num"]: r["nk"] for r in rows}


def _build_csv(qs: dict[str, list[str]]) -> bytes:
    rows = _query_messages(qs)
    detailed = qs.get("detail", ["0"])[0] in ("1", "true", "yes")
    buf = io.StringIO()
    writer = csv.writer(buf)
    if detailed:
        writer.writerow(["房间号", "主播", "时间", "分类", "用户", "内容", "礼物", "数量"])
        for r in rows:
            writer.writerow(
                [
                    r["room_num"], r["nickname"], r["created_at"],
                    _METHOD_CATEGORY.get(r["method"], ("other", "其他"))[1],
                    r["user_name"], r["content"], r["gift_name"], r["gift_count"],
                ]
            )
    else:
        writer.writerow(["房间号", "主播", "时间", "用户", "内容"])
        for r in rows:
            writer.writerow(
                [r["room_num"], r["nickname"], r["created_at"], r["user_name"], r["content"]]
            )
    # ﻿ BOM 让 Excel 正确识别 UTF-8 中文
    return ("﻿" + buf.getvalue()).encode("utf-8")


def _build_transcripts_csv(qs: dict[str, list[str]]) -> bytes:
    """话术导出 CSV：房间号、主播、时间、起始秒、话术文本。可按 room 过滤。"""
    nick_map = _room_nick_map()
    where = ""
    params: list[Any] = []
    if "room" in qs and qs["room"][0]:
        rooms = [r for r in qs["room"][0].split(",") if r]
        where = f"WHERE room_num IN ({','.join('?' * len(rooms))})"
        params.extend(rooms)
    conn = _connect(readonly=True)
    try:
        rows = conn.execute(
            f"SELECT room_num, created_at, start_sec, text FROM transcripts "
            f"{where} ORDER BY room_num, id",
            params,
        ).fetchall()
    finally:
        conn.close()
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["房间号", "主播", "时间", "起始秒", "话术文本"])
    for r in rows:
        writer.writerow(
            [r["room_num"], nick_map.get(r["room_num"], ""), r["created_at"],
             r["start_sec"], r["text"]]
        )
    return ("﻿" + buf.getvalue()).encode("utf-8")


_SHEET_BAD = str.maketrans({c: "_" for c in r"[]:*?/\\"})


def _sheet_title(nickname: str, room_num: str, used: set[str]) -> str:
    """生成合法、唯一的工作表名（Excel 限制 31 字符、禁用 []:*?/\\）。"""
    base = (nickname or room_num or "未知").translate(_SHEET_BAD).strip() or room_num
    base = base[:28]  # 预留 ~N 去重后缀
    title = base or room_num
    suffix = 1
    while title in used or not title:
        suffix += 1
        title = f"{base[:26]}~{suffix}"
    used.add(title)
    return title


def _build_xlsx(qs: dict[str, list[str]]) -> bytes:
    """多工作表导出，便于按主播打分：
    总览(弹幕/礼物/进入/话术段数·字数) + 每主播弹幕表 + 每主播·话术表。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    rows = _query_messages(qs)
    detailed = qs.get("detail", ["0"])[0] in ("1", "true", "yes")
    tr_counts = _transcript_counts()
    nick_map = _room_nick_map()

    # 按房间分组（保持查询顺序：room_num, id）
    grouped: dict[str, dict[str, Any]] = {}
    for r in rows:
        g = grouped.setdefault(r["room_num"], {"nickname": r["nickname"], "rows": []})
        if r["nickname"] and not g["nickname"]:
            g["nickname"] = r["nickname"]
        g["rows"].append(r)

    # 房间全集：有弹幕的 + 只有话术的(也要导出，便于打分)
    room_order: list[str] = list(grouped.keys())
    for rn in tr_counts:
        if rn not in grouped:
            room_order.append(rn)

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F7F52")
    title_font = Font(bold=True, size=12)

    def _style_header(ws_row) -> None:
        for cell in ws_row:
            cell.font = header_font
            cell.fill = header_fill

    msg_headers = (
        ["时间", "分类", "用户", "内容", "礼物", "数量"]
        if detailed
        else ["时间", "用户", "内容"]
    )

    # 总览表（打分面板）
    overview = wb.active
    overview.title = "总览"
    overview.append(["直播号", "主播", "本表消息数", "话术段数", "话术字数"])
    _style_header(overview[1])

    used_titles: set[str] = {"总览"}
    for room_num in room_order:
        g = grouped.get(room_num)
        nickname = (g["nickname"] if g else "") or nick_map.get(room_num, "")
        msg_n = len(g["rows"]) if g else 0
        tr_seg, tr_chars = tr_counts.get(room_num, (0, 0))
        overview.append([room_num, nickname or "-", msg_n, tr_seg, tr_chars])

        # 每主播弹幕表
        if g:
            ws = wb.create_sheet(_sheet_title(nickname, room_num, used_titles))
            ws.append([f"直播号 {room_num}", nickname or ""])
            ws["A1"].font = title_font
            ws.append(msg_headers)
            _style_header(ws[2])
            for r in g["rows"]:
                if detailed:
                    ws.append(
                        [
                            r["created_at"],
                            _METHOD_CATEGORY.get(r["method"], ("other", "其他"))[1],
                            r["user_name"], r["content"], r["gift_name"], r["gift_count"],
                        ]
                    )
                else:
                    ws.append([r["created_at"], r["user_name"], r["content"]])
            widths = [18, 8, 14, 60, 12, 6] if detailed else [18, 14, 60]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[chr(64 + i)].width = w

        # 每主播·话术表
        if tr_counts.get(room_num, (0, 0))[0]:
            base = (nickname or room_num)[:24]
            tws = wb.create_sheet(_sheet_title(f"{base}·话术", room_num, used_titles))
            tws.append([f"直播号 {room_num}", nickname or "", "话术"])
            tws["A1"].font = title_font
            tws.append(["时间", "起始秒", "话术文本"])
            _style_header(tws[2])
            for tr in _transcripts_for(room_num):
                tws.append([tr["created_at"], tr["start_sec"], tr["text"]])
            for i, w in enumerate([18, 8, 90], start=1):
                tws.column_dimensions[chr(64 + i)].width = w

    for col, w in (("A", 16), ("B", 22), ("C", 12), ("D", 10), ("E", 10)):
        overview.column_dimensions[col].width = w

    if not room_order:
        ws = wb.create_sheet("暂无数据")
        ws.append(["暂无数据，导入直播号并开始托管后刷新。"])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_index() -> bytes:
    rooms = _rooms_summary()
    tr_counts = _transcript_counts()
    lines = [
        "<!doctype html><html lang='zh'><head><meta charset='utf-8'>",
        "<title>弹幕导出</title>",
        "<style>body{font-family:system-ui;margin:24px;color:#18211d;background:#f4f6f5}"
        "h1{font-size:20px}table{border-collapse:collapse;width:100%;background:#fff;max-width:1100px}"
        "th,td{border:1px solid #dde5e0;padding:8px 10px;text-align:left;font-size:14px}"
        "th{background:#eef3f0}a{color:#1f7f52}.top a{margin-right:12px}"
        ".danger{margin-top:18px;padding:12px;background:#fff;border:1px solid #efcaca;max-width:1100px}"
        ".danger button{margin-right:8px;border:1px solid #b54848;background:#b54848;color:#fff;"
        "border-radius:6px;height:32px;padding:0 12px;cursor:pointer}"
        ".danger button.secondary{background:#fff;color:#b54848}</style></head><body>",
        "<h1>弹幕导出控制台</h1>",
        "<p class='top'>"
        "<a href='/export/messages.xlsx?category=chat'><b>导出Excel(弹幕+话术·每主播分表)</b></a>"
        "<a href='/export/messages.csv?category=chat'>聊天弹幕(CSV)</a>"
        "<a href='/export/messages.csv'>全部消息(CSV)</a>"
        "<a href='/export/transcripts.csv'>全部话术(CSV)</a>"
        "</p>",
        "<div class='danger'><b>数据清理</b>"
        "<p>清理后不可恢复。建议导出备份后再操作。</p>"
        "<form method='post' action='/admin/clear' "
        "onsubmit=\"return confirm('确认清除全部弹幕和话术数据？此操作不可恢复。')\">"
        "<input type='hidden' name='scope' value='all'>"
        "<button type='submit'>清除全部数据</button></form>"
        "<form method='post' action='/admin/clear' "
        "onsubmit=\"return confirm('确认只清除弹幕数据？')\">"
        "<input type='hidden' name='scope' value='messages'>"
        "<button class='secondary' type='submit'>只清弹幕</button></form>"
        "<form method='post' action='/admin/clear' "
        "onsubmit=\"return confirm('确认只清除话术数据？')\">"
        "<input type='hidden' name='scope' value='transcripts'>"
        "<button class='secondary' type='submit'>只清话术</button></form></div>",
        "<table><thead><tr><th>直播号</th><th>主播</th><th>标题</th>"
        "<th>聊天</th><th>礼物</th><th>进入</th><th>话术段</th><th>总计</th>"
        "<th>最近</th><th>导出</th></tr></thead><tbody>",
    ]
    for r in rooms:
        last = (
            datetime.fromtimestamp(r["last_seen"]).strftime("%m-%d %H:%M")
            if r["last_seen"]
            else "-"
        )
        rn = r["room_num"]
        tr_seg = tr_counts.get(rn, (0, 0))[0]
        lines.append(
            f"<tr><td>{rn}</td><td>{r['nickname'] or '-'}</td><td>{r['title'] or '-'}</td>"
            f"<td>{r['chat'] or 0}</td><td>{r['gift'] or 0}</td><td>{r['member'] or 0}</td>"
            f"<td>{tr_seg}</td><td>{r['total'] or 0}</td><td>{last}</td>"
            f"<td><a href='/export/messages.xlsx?category=chat&room={rn}'>Excel</a> · "
            f"<a href='/export/messages.csv?category=chat&room={rn}'>聊天</a> · "
            f"<a href='/export/transcripts.csv?room={rn}'>话术</a></td></tr>"
        )
    if not rooms:
        lines.append("<tr><td colspan='10'>暂无数据，导入直播号并开始托管后刷新。</td></tr>")
    lines.append("</tbody></table></body></html>")
    return "".join(lines).encode("utf-8")


class ExportHandler(BaseHTTPRequestHandler):
    def log_message(self, *args: Any) -> None:  # 静默默认访问日志
        pass

    def _send(self, body: bytes, content_type: str, filename: str | None = None) -> None:
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Access-Control-Allow-Origin", "*")
        if filename:
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        qs = parse_qs(parsed.query)
        try:
            if parsed.path in ("/", "/index.html"):
                self._send(_build_index(), "text/html; charset=utf-8")
            elif parsed.path == "/export/messages.csv":
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                self._send(_build_csv(qs), "text/csv; charset=utf-8", f"danmu_{stamp}.csv")
            elif parsed.path == "/export/messages.xlsx":
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                self._send(
                    _build_xlsx(qs),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    f"live_watch_{stamp}.xlsx",
                )
            elif parsed.path == "/export/transcripts.csv":
                stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                self._send(
                    _build_transcripts_csv(qs), "text/csv; charset=utf-8",
                    f"huashu_{stamp}.csv",
                )
            else:
                self.send_error(404, "Not Found")
        except Exception:  # noqa: BLE001
            logger.exception("导出请求失败 path=%s", self.path)
            self.send_error(500, "Export Error")

    def do_POST(self) -> None:  # noqa: N802
        parsed = urlparse(self.path)
        if parsed.path != "/admin/clear":
            self.send_error(404, "Not Found")
            return
        length = int(self.headers.get("Content-Length") or "0")
        body = self.rfile.read(length).decode("utf-8", errors="ignore")
        qs = parse_qs(body)
        scope = qs.get("scope", [""])[0]
        try:
            clear_data(scope)
            self.send_response(303)
            self.send_header("Location", "/")
            self.end_headers()
        except ValueError as exc:
            self.send_error(400, str(exc))
        except Exception:  # noqa: BLE001
            logger.exception("清理数据失败 scope=%s", scope)
            self.send_error(500, "Clear Error")


def clear_data(scope: str) -> None:
    if scope not in {"all", "messages", "transcripts"}:
        raise ValueError("Invalid clear scope")
    assert _write_conn is not None
    with _write_lock:
        if scope in {"all", "messages"}:
            _write_conn.execute("DELETE FROM messages")
            _write_conn.execute("UPDATE rooms SET msg_count = 0")
        if scope in {"all", "transcripts"}:
            _write_conn.execute("DELETE FROM transcripts")
        if scope == "all":
            _write_conn.execute("DELETE FROM rooms")
        _write_conn.commit()
    logger.warning("数据已清理 scope=%s", scope)


def start_http_server() -> ThreadingHTTPServer:
    httpd = ThreadingHTTPServer((HTTP_HOST, HTTP_PORT), ExportHandler)
    thread = threading.Thread(target=httpd.serve_forever, name="http-export", daemon=True)
    thread.start()
    logger.info("导出服务 => http://%s:%d", HTTP_HOST, HTTP_PORT)
    return httpd


# --------------------------------------------------------------------------- #
# 启动
# --------------------------------------------------------------------------- #
async def main() -> None:
    global _audio
    init_db()
    start_http_server()
    _audio = AudioManager(save_transcript)
    await _audio.warmup()
    try:
        async with websockets.serve(relay_handler, WS_HOST, WS_PORT):
            logger.info("中继服务 => ws://%s:%d", WS_HOST, WS_PORT)
            logger.info("等待前端 BatchView 转发弹幕+话术采集… (Ctrl+C 停止)")
            await asyncio.Future()  # run forever
    finally:
        if _audio is not None:
            await _audio.shutdown()


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        logger.info("已停止")

import json
import sqlite3
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook

from pipeline import export


class MonitoredRoomIdsTests(unittest.TestCase):
    def test_supports_legacy_string_room_list(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rooms.json"
            path.write_text(json.dumps(["123", "456"]), encoding="utf-8")
            with patch("pipeline.export.ROOMS_JSON", path):
                self.assertEqual(export.monitored_room_ids(), ["123", "456"])

    def test_extracts_rid_from_metadata_objects(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rooms.json"
            path.write_text(
                json.dumps(
                    [
                        {"rid": "123", "anchor_name": "主播", "avatar_url": "https://example.test/a.jpg"},
                        {"rid": "456", "source_url": "https://live.douyin.com/456?x=1"},
                    ],
                    ensure_ascii=False,
                ),
                encoding="utf-8",
            )
            with patch("pipeline.export.ROOMS_JSON", path):
                self.assertEqual(export.monitored_room_ids(), ["123", "456"])

    def test_ignores_invalid_entries_instead_of_using_dict_as_filename(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rooms.json"
            path.write_text(json.dumps([{"anchor_name": "missing"}, None, "not-a-room"]), encoding="utf-8")
            with patch("pipeline.export.ROOMS_JSON", path):
                self.assertEqual(export.monitored_room_ids(), [])

    def test_single_room_export_uses_safe_anchor_nickname(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "rooms.json"
            path.write_text(
                json.dumps([{"rid": "123", "anchor_name": "主播:A/B?"}], ensure_ascii=False),
                encoding="utf-8",
            )
            with (
                patch("pipeline.export.ROOMS_JSON", path),
                patch("pipeline.export.load_room_meta", return_value={}),
            ):
                self.assertEqual(export.room_export_filename("123"), "主播_A_B_.xlsx")

    def test_single_room_export_falls_back_to_live_id(self) -> None:
        with (
            patch("pipeline.export.configured_room_meta", return_value={}),
            patch("pipeline.export.load_room_meta", return_value={}),
        ):
            self.assertEqual(export.room_export_filename("123"), "123.xlsx")

    def test_duplicate_nicknames_append_live_id(self) -> None:
        class Bundle:
            def __init__(self, rid, nickname):
                self.rid = rid
                self.nickname = nickname

        stems = export.unique_bundle_stems([Bundle("123", "同名主播"), Bundle("456", "同名主播")])

        self.assertEqual(stems, {"123": "同名主播_123", "456": "同名主播_456"})

    def test_data_room_summaries_uses_aggregate_counts_not_full_bundle(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            transcript_db = root / "transcripts.db"
            events_db = root / "events.db"
            audio_dir = root / "audio"
            room_dir = audio_dir / "123"
            room_dir.mkdir(parents=True)
            (room_dir / "seq00001.mp3").write_bytes(b"a" * 10)
            (room_dir / "seq00002.mp3").write_bytes(b"b" * 20)

            tc = sqlite3.connect(transcript_db)
            tc.execute("CREATE TABLE transcripts(room_id TEXT)")
            tc.executemany("INSERT INTO transcripts(room_id) VALUES (?)", [("123",), ("123",)])
            tc.commit()
            tc.close()

            ec = sqlite3.connect(events_db)
            ec.execute("CREATE TABLE events(live_id TEXT, event_type TEXT, content TEXT)")
            ec.executemany(
                "INSERT INTO events(live_id, event_type, content) VALUES (?, ?, ?)",
                [
                    ("123", "chat", "hello"),
                    ("123", "chat", "world"),
                    ("123", "member", "in"),
                    ("123", "stat", "current=5;total_pv=99"),
                ],
            )
            ec.commit()
            ec.close()

            with (
                patch.object(export.config, "DB_PATH", transcript_db),
                patch.object(export.config, "EVENTS_DB", events_db),
                patch.object(export.config, "AUDIO_DIR", audio_dir),
                patch.object(export, "room_display_names", return_value={"123": "测试主播"}),
                patch.object(export, "build_bundle", side_effect=AssertionError("full bundle should not load")),
            ):
                summaries = export.data_room_summaries()

        self.assertEqual(len(summaries), 1)
        self.assertEqual(summaries[0]["rid"], "123")
        self.assertEqual(summaries[0]["nickname"], "测试主播")
        self.assertEqual(summaries[0]["transcripts"], 2)
        self.assertEqual(summaries[0]["events"], 4)
        self.assertEqual(summaries[0]["chats"], 2)
        self.assertEqual(summaries[0]["stats"], 1)
        self.assertEqual(summaries[0]["audio_files"], 2)
        self.assertEqual(summaries[0]["audio_bytes"], 30)

    def test_summary_xlsx_includes_license_watermark_when_policy_enabled(self) -> None:
        bundle = export.RoomBundle(
            rid="123",
            nickname="测试主播",
            transcripts=[],
            timeline=[],
            chats=[],
            stats=[],
            event_counts={},
        )

        class Status:
            ok = True
            mode = "licensed"
            payload = {"license_id": "lic-001", "activation_id": "act-001"}

        with (
            patch.object(export, "load_chats_ts", return_value=[]),
            patch.object(export.license_manager, "current_policy", return_value={"export_watermark": True}),
            patch.object(export.license_manager, "current_status", return_value=Status()),
            patch.object(export.license_manager, "current_device_hash", return_value="abcdef1234567890"),
            patch.object(export.config, "LICENSE_APP_VERSION", "1.2.3"),
        ):
            wb = load_workbook(BytesIO(export.summary_xlsx_bytes([bundle])))

        self.assertIn("授权水印", wb.sheetnames)
        rows = {row[0].value: row[1].value for row in wb["授权水印"].iter_rows(min_row=2, max_col=2)}
        self.assertEqual(rows["授权ID"], "lic-001")
        self.assertEqual(rows["设备指纹"], "abcdef123456")
        self.assertEqual(rows["授权状态"], "licensed")

    def test_summary_xlsx_skips_license_watermark_when_policy_disabled(self) -> None:
        bundle = export.RoomBundle(
            rid="123",
            nickname="测试主播",
            transcripts=[],
            timeline=[],
            chats=[],
            stats=[],
            event_counts={},
        )

        with (
            patch.object(export, "load_chats_ts", return_value=[]),
            patch.object(export.license_manager, "current_policy", return_value={"export_watermark": False}),
        ):
            wb = load_workbook(BytesIO(export.summary_xlsx_bytes([bundle])))

        self.assertNotIn("授权水印", wb.sheetnames)


if __name__ == "__main__":
    unittest.main()
